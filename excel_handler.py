"""Excel export — generates Claim Master Sheet."""

import tempfile
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import PageSetupProperties


def build_workbook(rows, output_path=None):
    """Build an xlsx workbook from invoice rows. Returns the file path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        "BRANCH", "SUPPLIERS NAME", "INVOICE NO.", "INVOICE DATE",
        "CATEGORY", "DESCRIPTION", "AMOUNT (RM)", "CLAIM DATE",
    ]

    # Header style
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    hdr_fill = PatternFill(start_color="1E2A3A", end_color="1E2A3A", fill_type="solid")
    hdr_border = Border(bottom=Side(border_style="medium", color="4F6EF7"))

    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = hdr_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Data rows
    alt_fill = PatternFill(start_color="F8FAFB", end_color="F8FAFB", fill_type="solid")
    num_fmt = '#,##0.00'
    wrap_align = Alignment(wrap_text=True, vertical="center")

    for idx, row in enumerate(rows, 2):
        amt_raw = row.get("amount", "") or ""
        try:
            amt = float(str(amt_raw).replace(",", "").strip())
        except (ValueError, TypeError):
            amt = amt_raw

        ws.append([
            row.get("branch", ""),
            row.get("supplierName", ""),
            row.get("invoiceNo", ""),
            row.get("invoiceDate", ""),
            row.get("category", ""),
            row.get("description", ""),
            amt,
            row.get("claimDate", ""),
        ])

        if idx % 2 == 0:
            for cell in ws[idx]:
                cell.fill = alt_fill

        amt_cell = ws.cell(row=idx, column=7)
        amt_cell.number_format = num_fmt
        amt_cell.alignment = Alignment(horizontal="right", vertical="center")

        # Wrap text for long content columns (B, C, F)
        for col in (2, 3, 6):
            ws.cell(row=idx, column=col).alignment = wrap_align

        # Center-align date columns (D=Invoice Date, H=Claim Date)
        for col in (4, 8):
            ws.cell(row=idx, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # TOTAL row below data
    last_row = max(len(rows) + 1, 2)
    total_row = last_row + 1
    total_border = Border(top=Side(border_style="medium", color="4F6EF7"))

    ws.cell(row=total_row, column=6, value="TOTAL").font = Font(bold=True, name="Calibri", size=10)
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=total_row, column=6).border = total_border

    ws.cell(row=total_row, column=7).value = f"=SUM(G2:G{last_row})"
    ws.cell(row=total_row, column=7).number_format = num_fmt
    ws.cell(row=total_row, column=7).font = Font(bold=True, name="Calibri", size=10)
    ws.cell(row=total_row, column=7).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=total_row, column=7).border = total_border

    # Column widths (wider for long content)
    widths = {"A": 8, "B": 42, "C": 28, "D": 14, "E": 16, "F": 50, "G": 14, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"

    # Print settings: landscape, fit to 1 page wide, repeat header row
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:1"

    if output_path:
        wb.save(output_path)
        return output_path

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


def export_filename():
    return f"Claim_Master_Sheet_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
