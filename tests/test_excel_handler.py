"""Tests for excel_handler.py — workbook generation and SUBTOTAL formula."""

import os
import openpyxl
import pytest

from excel_handler import build_workbook, export_filename


class TestBuildWorkbook:
    def _make_rows(self, n):
        return [
            {
                "branch": "HQ",
                "supplierName": f"Supplier {i}",
                "invoiceNo": f"INV-{i:04d}",
                "invoiceDate": "01/01/2026",
                "category": "Advertisement",
                "description": "Google ads",
                "amount": str(100 + i),
                "claimDate": "15/01/2026",
            }
            for i in range(n)
        ]

    def test_creates_file(self, tmp_dir):
        path = os.path.join(tmp_dir, "test.xlsx")
        result = build_workbook(self._make_rows(3), output_path=path)
        assert os.path.exists(result)

    def test_headers_present(self, tmp_dir):
        path = os.path.join(tmp_dir, "test.xlsx")
        build_workbook(self._make_rows(1), output_path=path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        assert headers == [
            "BRANCH", "SUPPLIERS NAME", "INVOICE NO.", "INVOICE DATE",
            "CATEGORY", "DESCRIPTION", "AMOUNT (RM)", "CLAIM DATE",
        ]

    def test_data_rows_written(self, tmp_dir):
        path = os.path.join(tmp_dir, "test.xlsx")
        rows = self._make_rows(5)
        build_workbook(rows, output_path=path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.max_row == 7  # 1 header + 5 data + 1 total

    def test_amount_parsed_as_number(self, tmp_dir):
        path = os.path.join(tmp_dir, "test.xlsx")
        build_workbook([{
            "branch": "HQ", "supplierName": "Test", "invoiceNo": "1",
            "invoiceDate": "", "category": "", "description": "",
            "amount": "1,234.56", "claimDate": "",
        }], output_path=path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(row=2, column=7).value == 1234.56

    def test_total_row_below_data(self, tmp_dir):
        """TOTAL row should be placed below the last data row."""
        path = os.path.join(tmp_dir, "test.xlsx")
        build_workbook(self._make_rows(10), output_path=path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Total should be in row 12 (1 header + 10 data + 1 total)
        assert ws.cell(row=12, column=6).value == "TOTAL"
        assert ws.cell(row=12, column=7).value == "=SUM(G2:G11)"

    def test_total_row_single_row(self, tmp_dir):
        path = os.path.join(tmp_dir, "test.xlsx")
        build_workbook(self._make_rows(1), output_path=path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(row=3, column=6).value == "TOTAL"
        assert ws.cell(row=3, column=7).value == "=SUM(G2:G2)"

    def test_wider_columns(self, tmp_dir):
        """Key columns should be wider for print readability."""
        path = os.path.join(tmp_dir, "test.xlsx")
        build_workbook(self._make_rows(1), output_path=path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.column_dimensions["B"].width == 42  # Supplier Name
        assert ws.column_dimensions["C"].width == 28  # Invoice No
        assert ws.column_dimensions["F"].width == 50  # Description

    def test_print_settings(self, tmp_dir):
        """Landscape, fit to 1 page wide, repeat header row."""
        path = os.path.join(tmp_dir, "test.xlsx")
        build_workbook(self._make_rows(1), output_path=path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.page_setup.orientation == "landscape"
        assert ws.print_title_rows == "$1:$1"

    def test_empty_rows(self, tmp_dir):
        path = os.path.join(tmp_dir, "test.xlsx")
        build_workbook([], output_path=path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.max_row == 3  # header + empty data area + total row

    def test_temp_file_when_no_path(self):
        path = build_workbook([{
            "branch": "", "supplierName": "", "invoiceNo": "",
            "invoiceDate": "", "category": "", "description": "",
            "amount": "50", "claimDate": "",
        }])
        assert path.endswith(".xlsx")
        assert os.path.exists(path)
        os.unlink(path)


class TestExportFilename:
    def test_format(self):
        name = export_filename()
        assert name.startswith("Claim_Master_Sheet_")
        assert name.endswith(".xlsx")
